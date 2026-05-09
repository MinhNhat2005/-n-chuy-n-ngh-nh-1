import os
from io import BytesIO
from flask import Blueprint, jsonify, request, send_file
from db import get_db

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment

attendance_history_bp = Blueprint("attendance_history", __name__)
attendance_export_bp = Blueprint("attendance_export", __name__)


# =========================================================
# 📜 LẤY LỊCH SỬ ĐIỂM DANH
# =========================================================
@attendance_history_bp.route("/attendance/history", methods=["GET"])
def get_history():

    try:
        db = get_db()
        cursor = db.cursor()

        search = request.args.get("search", "")
        class_id = request.args.get("classId", "all")

        query = """
            SELECT 
                id,
                student_id,
                student_name,
                class_id,
                session,
                time,
                image
            FROM attendance
            WHERE 1=1
        """

        params = []

        # 🔍 SEARCH
        if search:
            query += """
                AND (
                    student_name LIKE %s
                    OR student_id LIKE %s
                )
            """

            params.append(f"%{search}%")
            params.append(f"%{search}%")

        # 🎯 FILTER CLASS
        if class_id != "all":
            query += " AND class_id = %s"
            params.append(class_id)

        query += " ORDER BY time DESC"

        cursor.execute(query, params)

        rows = cursor.fetchall()

        result = []

        for row in rows:

            result.append({
                "id": row[0],
                "student_id": row[1],
                "student_name": row[2],
                "class_name": row[3],
                "session": row[4],
                "time": row[5].strftime("%Y-%m-%d %H:%M:%S"),
                "image": row[6]
            })

        cursor.close()
        db.close()

        return jsonify(result), 200

    except Exception as e:
        print("ERROR:", e)

        return jsonify({
            "error": "Server error"
        }), 500


# =========================================================
# 🗑️ XÓA LỊCH SỬ ĐIỂM DANH
# =========================================================
@attendance_history_bp.route("/attendance/<int:id>", methods=["DELETE"])
def delete_attendance(id):

    try:
        db = get_db()
        cursor = db.cursor()

        # 🔥 LẤY ẢNH
        cursor.execute(
            "SELECT image FROM attendance WHERE id = %s",
            (id,)
        )

        row = cursor.fetchone()

        if not row:
            return jsonify({
                "error": "Record not found"
            }), 404

        image_path = row[0]

        # 🔥 XÓA FILE ẢNH
        if image_path:

            try:
                if os.path.exists(image_path):
                    os.remove(image_path)

                    print("Deleted image:", image_path)

            except Exception as e:
                print("Delete image error:", e)

        # 🔥 XÓA DB
        cursor.execute(
            "DELETE FROM attendance WHERE id = %s",
            (id,)
        )

        db.commit()

        cursor.close()
        db.close()

        return jsonify({
            "message": "Deleted successfully"
        }), 200

    except Exception as e:
        print("ERROR:", e)

        return jsonify({
            "error": "Server error"
        }), 500


# =========================================================
# 📤 EXPORT EXCEL
# =========================================================
@attendance_export_bp.route("/attendance/export", methods=["GET"])
def export_excel():

    try:
        db = get_db()
        cursor = db.cursor()

        search = request.args.get("search", "")
        class_id = request.args.get("classId", "all")

        query = """
            SELECT
                student_id,
                student_name,
                class_id,
                session,
                time,
                image
            FROM attendance
            WHERE 1=1
        """

        params = []

        # 🔍 SEARCH
        if search:
            query += """
                AND (
                    student_name LIKE %s
                    OR student_id LIKE %s
                )
            """

            params.append(f"%{search}%")
            params.append(f"%{search}%")

        # 🎯 FILTER CLASS
        if class_id != "all":
            query += " AND class_id = %s"
            params.append(class_id)

        query += " ORDER BY time DESC"

        cursor.execute(query, params)

        rows = cursor.fetchall()

        # =====================================================
        # 📗 TẠO FILE EXCEL
        # =====================================================
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # =====================================================
        # 🏷️ TITLE
        # =====================================================
        if class_id == "all":
            title = "DANH SÁCH ĐIỂM DANH SINH VIÊN (TẤT CẢ LỚP)"
        else:
            title = f"DANH SÁCH ĐIỂM DANH LỚP {class_id.upper()}"

        ws.merge_cells("A1:F1")

        ws["A1"] = title

        ws["A1"].font = Font(
            size=14,
            bold=True
        )

        ws["A1"].alignment = Alignment(
            horizontal="center"
        )

        # =====================================================
        # 📌 HEADER
        # =====================================================
        headers = [
            "MSSV",
            "Họ tên",
            "Lớp",
            "Buổi",
            "Thời gian",
            "Ảnh"
        ]

        ws.append(headers)

        # =====================================================
        # 📏 WIDTH
        # =====================================================
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 20

        # =====================================================
        # 📥 DATA
        # =====================================================
        row_index = 3

        for row in rows:

            student_id = row[0]
            name = row[1]
            class_name = row[2]
            session = row[3]
            time = row[4]
            image_path = row[5]

            # TEXT
            ws.cell(row=row_index, column=1, value=student_id)
            ws.cell(row=row_index, column=2, value=name)
            ws.cell(row=row_index, column=3, value=class_name)
            ws.cell(row=row_index, column=4, value=session)
            ws.cell(row=row_index, column=5, value=str(time))

            # 🖼️ IMAGE
            if image_path and os.path.exists(image_path):

                try:
                    img = ExcelImage(image_path)

                    img.width = 80
                    img.height = 80

                    ws.add_image(img, f"F{row_index}")

                    ws.row_dimensions[row_index].height = 65

                except Exception as e:
                    print("IMAGE ERROR:", e)

            row_index += 1

        # =====================================================
        # 💾 EXPORT
        # =====================================================
        output = BytesIO()

        wb.save(output)

        output.seek(0)

        cursor.close()
        db.close()

        return send_file(
            output,
            as_attachment=True,
            download_name="attendance.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:

        print("EXPORT ERROR:", e)

        return jsonify({
            "error": "Export failed"
        }), 500